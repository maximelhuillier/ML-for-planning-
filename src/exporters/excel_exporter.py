"""
Excel Exporter for Delay Analysis Results
"""
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Optional
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..analyzers.base_analyzer import DelayAnalysisResult
from ..utils.schedule_utils import Schedule


class ExcelExporter:
    """Export delay analysis results to Excel format"""

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        # Define color scheme
        self.colors = {
            'header': 'FF366092',  # Dark blue
            'subheader': 'FF4472C4',  # Medium blue
            'critical': 'FFFF0000',  # Red
            'warning': 'FFFFC000',  # Orange
            'success': 'FF70AD47',  # Green
            'light_gray': 'FFD9D9D9',
        }

    def export(self, result: DelayAnalysisResult, output_path: str,
               include_charts: bool = True) -> str:
        """
        Export delay analysis result to Excel

        Args:
            result: DelayAnalysisResult object
            output_path: Path to save Excel file
            include_charts: Whether to include charts

        Returns:
            Path to saved file
        """
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create sheets
        self._create_summary_sheet(wb, result)
        self._create_detailed_sheet(wb, result)
        self._create_by_cause_sheet(wb, result)

        if include_charts:
            self._create_charts_sheet(wb, result)

        # Save workbook
        wb.save(output_path)
        return output_path

    def export_to_bytes(self, result: DelayAnalysisResult,
                       include_charts: bool = True) -> bytes:
        """
        Export to Excel format in memory (for downloads)

        Args:
            result: DelayAnalysisResult object
            include_charts: Whether to include charts

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        self._create_summary_sheet(wb, result)
        self._create_detailed_sheet(wb, result)
        self._create_by_cause_sheet(wb, result)

        if include_charts:
            self._create_charts_sheet(wb, result)

        # Save to bytes
        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        return virtual_workbook.getvalue()

    def _create_summary_sheet(self, wb: Workbook, result: DelayAnalysisResult):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "Delay Analysis Report"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'],
                                    end_color=self.colors['header'],
                                    fill_type='solid')
        ws.merge_cells('A1:D1')

        # Method and date
        row = 3
        ws[f'A{row}'] = "Analysis Method:"
        ws[f'B{row}'] = result.method_name
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Analysis Date:"
        ws[f'B{row}'] = result.analysis_date.strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)

        # Key metrics
        row += 2
        ws[f'A{row}'] = "KEY METRICS"
        ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['subheader'],
                                        end_color=self.colors['subheader'],
                                        fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')

        metrics = [
            ("Total Delay", f"{result.total_delay_days:.1f} days"),
            ("Critical Path Delay", f"{result.critical_delay_days:.1f} days"),
            ("Affected Activities", len(result.delays_by_activity)),
            ("Critical Activities Delayed", len([d for d in result.delays_by_activity if d['is_critical']])),
        ]

        for metric, value in metrics:
            row += 1
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)

        # Top delay causes
        if result.delays_by_cause:
            row += 2
            ws[f'A{row}'] = "TOP DELAY CAUSES"
            ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFFFF')
            ws[f'A{row}'].fill = PatternFill(start_color=self.colors['subheader'],
                                            end_color=self.colors['subheader'],
                                            fill_type='solid')
            ws.merge_cells(f'A{row}:D{row}')

            row += 1
            ws[f'A{row}'] = "Cause"
            ws[f'B{row}'] = "Days"
            ws[f'C{row}'] = "% of Total"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = Font(bold=True)
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.colors['light_gray'],
                                                     end_color=self.colors['light_gray'],
                                                     fill_type='solid')

            sorted_causes = sorted(result.delays_by_cause.items(),
                                 key=lambda x: x[1], reverse=True)[:5]

            for cause, days in sorted_causes:
                row += 1
                ws[f'A{row}'] = cause
                ws[f'B{row}'] = f"{days:.1f}"
                pct = (days / result.total_delay_days * 100) if result.total_delay_days > 0 else 0
                ws[f'C{row}'] = f"{pct:.1f}%"

        # Recommendations
        if result.recommendations:
            row += 2
            ws[f'A{row}'] = "RECOMMENDATIONS"
            ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFFFF')
            ws[f'A{row}'].fill = PatternFill(start_color=self.colors['subheader'],
                                            end_color=self.colors['subheader'],
                                            fill_type='solid')
            ws.merge_cells(f'A{row}:D{row}')

            for rec in result.recommendations:
                row += 1
                ws[f'A{row}'] = f"• {rec}"
                ws.merge_cells(f'A{row}:D{row}')
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_detailed_sheet(self, wb: Workbook, result: DelayAnalysisResult):
        """Create detailed activities sheet"""
        ws = wb.create_sheet("Detailed Analysis")

        # Title
        ws['A1'] = "Detailed Delay Analysis"
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'],
                                    end_color=self.colors['header'],
                                    fill_type='solid')

        # Create DataFrame
        if result.detailed_report is not None and not result.detailed_report.empty:
            df = result.detailed_report
        else:
            df = pd.DataFrame(result.delays_by_activity)

        # Write data starting from row 3
        row_num = 3

        # Headers
        if not df.empty:
            for col_num, column in enumerate(df.columns, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = column.replace('_', ' ').title()
                cell.font = Font(bold=True, color='FFFFFFFF')
                cell.fill = PatternFill(start_color=self.colors['subheader'],
                                       end_color=self.colors['subheader'],
                                       fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # Data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), row_num + 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value

                    # Format dates
                    if isinstance(value, datetime):
                        cell.number_format = 'YYYY-MM-DD'

                    # Highlight critical activities
                    if df.columns[c_idx-1] == 'is_critical' and value:
                        for col in range(1, len(df.columns) + 1):
                            ws.cell(row=r_idx, column=col).fill = PatternFill(
                                start_color='FFFFE0E0',
                                end_color='FFFFE0E0',
                                fill_type='solid'
                            )

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(50, max(12, max_length + 2))
                ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A4'

    def _create_by_cause_sheet(self, wb: Workbook, result: DelayAnalysisResult):
        """Create delays by cause sheet"""
        ws = wb.create_sheet("By Cause")

        # Title
        ws['A1'] = "Delays by Cause"
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'],
                                    end_color=self.colors['header'],
                                    fill_type='solid')

        # Create summary DataFrame
        if result.delays_by_cause:
            data = []
            for cause, days in sorted(result.delays_by_cause.items(),
                                     key=lambda x: x[1], reverse=True):
                pct = (days / result.total_delay_days * 100) if result.total_delay_days > 0 else 0
                data.append({
                    'Cause': cause,
                    'Delay (Days)': days,
                    'Percentage': f"{pct:.1f}%",
                    'Impact': 'High' if pct > 20 else 'Medium' if pct > 10 else 'Low'
                })

            df = pd.DataFrame(data)

            # Write to sheet
            row_num = 3

            # Headers
            for col_num, column in enumerate(df.columns, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = column
                cell.font = Font(bold=True, color='FFFFFFFF')
                cell.fill = PatternFill(start_color=self.colors['subheader'],
                                       end_color=self.colors['subheader'],
                                       fill_type='solid')

            # Data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), row_num + 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value

                    # Color code impact
                    if c_idx == 4:  # Impact column
                        if value == 'High':
                            cell.fill = PatternFill(start_color='FFFF6B6B',
                                                   end_color='FFFF6B6B',
                                                   fill_type='solid')
                        elif value == 'Medium':
                            cell.fill = PatternFill(start_color='FFFFD93D',
                                                   end_color='FFFFD93D',
                                                   fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FF95E1D3',
                                                   end_color='FF95E1D3',
                                                   fill_type='solid')

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12

    def _create_charts_sheet(self, wb: Workbook, result: DelayAnalysisResult):
        """Create charts sheet"""
        ws = wb.create_sheet("Charts")

        ws['A1'] = "Delay Analysis Charts"
        ws['A1'].font = Font(size=14, bold=True)

        # Pie chart for delays by cause
        if result.delays_by_cause and len(result.delays_by_cause) > 0:
            # Create data for chart
            row = 3
            ws[f'A{row}'] = "Cause"
            ws[f'B{row}'] = "Days"

            sorted_causes = sorted(result.delays_by_cause.items(),
                                 key=lambda x: x[1], reverse=True)[:8]

            for cause, days in sorted_causes:
                row += 1
                ws[f'A{row}'] = cause
                ws[f'B{row}'] = days

            # Create pie chart
            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=4, max_row=row)
            data = Reference(ws, min_col=2, min_row=3, max_row=row)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Delays by Cause"
            pie.height = 12
            pie.width = 20

            ws.add_chart(pie, "D3")


def export_to_excel(result: DelayAnalysisResult, output_path: str,
                   include_charts: bool = True) -> str:
    """
    Convenience function to export results to Excel

    Args:
        result: DelayAnalysisResult object
        output_path: Path to save Excel file
        include_charts: Whether to include charts

    Returns:
        Path to saved file
    """
    exporter = ExcelExporter()
    return exporter.export(result, output_path, include_charts)
